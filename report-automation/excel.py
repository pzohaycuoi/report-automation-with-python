import logging
import openpyxl
import common


common.logger_config()


@common.log_function_call
def load_excel(template_file_path):
    """
    Load excel workbook in local host
    """
    # convert to real path
    real_path = common.convert_to_realpath(template_file_path)
    default_template_path = common.convert_to_realpath(
        '../template/Book1.xlsx')
    wb = openpyxl.load_workbook(filename=real_path or default_template_path)
    return wb


@common.log_function_call
def write_excel(workbook, data, worksheet, table_name):
    """
    Insert data from database to excel file with excel template
    """
    @common.log_function_call
    def _load_worksheet(workbook, worksheet):
        """
        Active the worksheet of the workbook
        """
        ws = workbook[worksheet]
        return ws

    @common.log_function_call
    def _cleanup_worksheet(worksheet):
        """
        Clean up the worksheet - preparation for inserting data into worksheet
        """
        start_row = 2
        max_row = worksheet.max_row
        worksheet.delete_rows(start_row, max_row)

    @common.log_function_call
    def _insert_data(worksheet, data):
        """
        Insert rows of data into worksheet
        """
        start_row = 2
        for i, row_data in enumerate(data, start=start_row):
            for j, value in enumerate(row_data):
                worksheet.cell(row=i, column=j+1, value=value)

    @common.log_function_call
    def _colnum_string(col_max):
        """
        Convert column integer to excel column character
        """
        string = ""
        while col_max > 0:
            col_max, remainder = divmod(col_max - 1, 26)
            string = chr(65 + remainder) + string
        return string

    @common.log_function_call
    def _update_table_ref(worksheet, table_name):
        try:
            table = worksheet.tables[table_name]
            max_col = _colnum_string(worksheet.max_column)
            max_row = worksheet.max_row

            # If max row is 1 then set to table ref to 2
            # For better visualization
            if max_row == 1:
                table.ref = f'A1:{max_col}2'
            else:
                table.ref = f'A1:{max_col}{max_row}'

        except KeyError:
            logging.warning('Worksheet %s does not contain table %s',
                            worksheet, table_name)

    ws = _load_worksheet(workbook, worksheet)
    _cleanup_worksheet(ws)
    _insert_data(ws, data)
    _update_table_ref(ws, table_name)


@common.log_function_call
def refresh_pivot_table(workbook):
    """
    Refresh all pivot table in the workbook
    """
    for worksheet in workbook:
        if worksheet._pivots != []:
            for pivot_tbl in worksheet._pivots:
                pivot_tbl.cache.refreshOnLoad = True


@common.log_function_call
def save_workbook(workbook, destination_file_path):
    """
    Save the workbook to new destination
    """
    # Get real path
    real_path = common.convert_to_realpath(destination_file_path)
    workbook.save(real_path)


if __name__ == '__main__':
    template_file = '../template/Book1.xlsx'
    destination_file = ('../template/Book2.xlsx')
    wb = load_excel(template_file)
    refresh_pivot_table(wb)
    save_workbook(wb, destination_file)
    print()
